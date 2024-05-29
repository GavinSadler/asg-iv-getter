import datetime
import os
from typing import List

import pandas as pd
from openpyxl import Workbook, load_workbook
from PySide6.QtCore import QObject, Slot
from PySide6.QtWidgets import QComboBox, QDoubleSpinBox, QFileDialog, QMessageBox, QWidget

from MainWindow import MainWindow
from Measurement import DatastreamMeasurementWorker, DatastreamMode, DatastreamParameters, MeasurementPoint, SweepMeasurementWorker, SweepParameters
from PlotParamsDialog import PlotParam
from SourceMeter import ConnectSMUWorker, Source, SourceMeter


class Controller(QObject):

    sourcemeters: List[SourceMeter]
    main_window: MainWindow

    ds_data_file_path: str
    sm_data_file_path: str

    sm_data: pd.DataFrame | None
    ds_data: pd.DataFrame | None

    sm_last_run: List[pd.DataFrame]
    ds_last_run: pd.DataFrame

    ms_plot_param_x: PlotParam
    ms_plot_param_y: PlotParam
    ds_plot_param_x: PlotParam
    ds_plot_param_y: PlotParam

    _widget_enabled_state: dict[QWidget, bool]

    def __init__(self):
        super().__init__()

        # Initialize fields
        self.sourcemeters = []
        self.sm_data = None
        self.ds_data = None
        self.sm_last_run = [initialize_data()]
        self.ds_last_run = initialize_data()
        self.ms_plot_param_x = PlotParam.smu_1_voltage
        self.ms_plot_param_y = PlotParam.smu_1_current
        self.ds_plot_param_x = PlotParam.time
        self.ds_plot_param_y = PlotParam.smu_1_current
        self.ds_data_file_path = ""
        self.sm_data_file_path = ""
        self._widget_enabled_state = {}

        # Create the user interface
        self.main_window = MainWindow()
        self.main_window.show()

        # Setup the plotes
        self.main_window.setup_plots(self.sm_last_run[-1], self.ds_last_run)

        # === SMU Connection Tab ===
        self.main_window.smu_search.clicked.connect(self.smu_search_clicked)
        self.main_window.smu_connection_list.itemSelectionChanged.connect(self.smu_connection_selection_changed)
        self.main_window.smu_disconnect.clicked.connect(self.smu_disconnect_clicked)
        self.main_window.smu_identify.clicked.connect(self.smu_identify_clicked)

        # === Sweep tab ===
        self.main_window.sm_measurement_run.clicked.connect(self.sm_run_measurement)
        self.main_window.sm_quick_measurement_run.clicked.connect(lambda: self.sm_run_measurement(quick_measurement=True))

        self.main_window.sm_save_data.clicked.connect(self.sm_save_last_run)

        self.main_window.sm_plot_settings.clicked.connect(self.main_window.sm_plot.show_plot_params_dialog)

        self.main_window.sm_data_path_button.clicked.connect(self.sm_choose_data_file_path)

        self.main_window.sm_voltage_1.clicked.connect(self.sm_update_sources)
        self.main_window.sm_current_1.clicked.connect(self.sm_update_sources)
        self.main_window.sm_sweep_1.clicked.connect(self.sm_update_sources)
        self.main_window.sm_voltage_2.clicked.connect(self.sm_update_sources)
        self.main_window.sm_current_2.clicked.connect(self.sm_update_sources)
        self.main_window.sm_sweep_2.clicked.connect(self.sm_update_sources)

        # === Datastream tab ===
        self.main_window.ds_stream.clicked.connect(self.ds_stream_clicked)

        self.main_window.ds_save_data.clicked.connect(self.ds_save_last_run)

        self.main_window.ds_plot_settings.clicked.connect(self.main_window.ds_plot.show_plot_params_dialog)

        self.main_window.ds_data_path_button.clicked.connect(self.ds_choose_data_file_path)

        self.main_window.ds_voltage_1.clicked.connect(self.ds_update_sources)
        self.main_window.ds_current_1.clicked.connect(self.ds_update_sources)
        self.main_window.ds_voltage_2.clicked.connect(self.ds_update_sources)
        self.main_window.ds_current_2.clicked.connect(self.ds_update_sources)

        self.main_window.ds_continuous.clicked.connect(self.ds_update_mode)
        self.main_window.ds_fixed_duration.clicked.connect(self.ds_update_mode)
        self.main_window.ds_fixed_num.clicked.connect(self.ds_update_mode)

        # Update source inputs to reflect real values
        self.sm_update_sources()
        self.ds_update_sources()

    @Slot()
    def smu_search_clicked(self):
        # Disconnect all currently connected SMUs
        self.disconnect_all_smus()

        # Create the search thread
        self._smu_search_thread = ConnectSMUWorker(self)
        self._smu_search_thread.progress_update.connect(self.main_window.smu_search_progress.setValue)
        self._smu_search_thread.status_update.connect(self.main_window.smu_search_log.appendPlainText)
        self._smu_search_thread.connections_made.connect(self.connect_smus)
        self._smu_search_thread.finished.connect(self.smu_search_finished)

        # Update the search button
        self.main_window.smu_search.setText("Cancel SMU Search")
        self.main_window.smu_search.clicked.disconnect()
        self.main_window.smu_search.clicked.connect(self._smu_search_thread.cancel_search)

        # Disable the connections list
        self.main_window.smu_connections.setDisabled(True)

        # Start the search
        self._smu_search_thread.start()

    @Slot()
    def smu_search_finished(self):
        # Update the search button
        self.main_window.smu_search.setText("Search for SMUs")
        self.main_window.smu_search.clicked.disconnect()
        self.main_window.smu_search.clicked.connect(self.smu_search_clicked)

        # Enable the connections list
        self.main_window.smu_connections.setDisabled(False)

        # Reset the progress bar
        self.main_window.smu_search_progress.reset()

    @Slot()
    def connect_smus(self, smus: List[SourceMeter]):
        # Disconnect any SMUs that may be connected
        self.disconnect_all_smus()
        self.sourcemeters = smus
        self.update_smu_uis()

    @Slot()
    def disconnect_all_smus(self):
        # Make sure to gracefully disconnect SMUs
        for smu in self.sourcemeters:
            smu.disconnect()

        self.sourcemeters = []

        self.update_smu_uis()

    def update_smu_uis(self):
        self.main_window.smu_connection_list.clear()

        # Add the SMU to the connections list
        for smu in self.sourcemeters:
            self.main_window.smu_connection_list.addItem(smu.serial_number)

        combo_box: QComboBox

        for combo_box in self.main_window.findChildren(QComboBox):

            if "smu_select" in combo_box.objectName():
                combo_box.clear()
                combo_box.addItem("Simulated")

                for smu in self.sourcemeters:
                    combo_box.addItem(smu.serial_number)

    def get_smu_from_serial(self, serial: str, allow_simulated=False):
        for smu in self.sourcemeters:
            if smu.serial_number == serial:
                return smu

        if allow_simulated:
            return SourceMeter()

    def get_selected_smu(self):

        if len(self.main_window.smu_connection_list.selectedItems()) == 0:
            return None

        serial = self.main_window.smu_connection_list.selectedItems()[0].text()

        return self.get_smu_from_serial(serial)

    @Slot()
    def smu_connection_selection_changed(self):
        smu = self.get_selected_smu()

        if smu:
            self.main_window.smu_disconnect.setDisabled(False)
            self.main_window.smu_identify.setDisabled(False)
        else:
            self.main_window.smu_disconnect.setDisabled(True)
            self.main_window.smu_identify.setDisabled(True)

    @Slot()
    def smu_disconnect_clicked(self):

        smu = self.get_selected_smu()

        if smu:
            smu.disconnect()
            self.sourcemeters.remove(smu)
            self.update_smu_uis()

    @Slot()
    def smu_identify_clicked(self):
        smu = self.get_selected_smu()

        if smu:
            smu.beep(1000, 0.5)

    def get_sweep_parameters(self, quick_measurement=False):

        # Identify which SMU will be used as the sweep, and which will be used as the constant
        # as well as what source mode each is in
        if self.main_window.sm_sweep_1.isChecked():
            sweep_souce = self.sm_get_smu_1_source()
            constant_source = self.sm_get_smu_2_source()
        else:
            sweep_souce = self.sm_get_smu_2_source()
            constant_source = self.sm_get_smu_1_source()

        sweep_start = self.main_window.sm_sweep_start.value()
        sweep_step = self.main_window.sm_sweep_step.value() if not quick_measurement else self.main_window.sm_quick_measurement_sweep_step.value()
        sweep_end = self.main_window.sm_sweep_end.value()
        sweep_compliance = self.main_window.sm_sweep_compliance.value()

        constant_output = self.main_window.sm_constant_output.value()
        constant_compliance = self.main_window.sm_constant_compliance.value()

        measurement_pause = self.main_window.sm_measurement_pause.value() if not quick_measurement else self.main_window.sm_quick_measurement_pause.value()
        sweep_pause = self.main_window.sm_sweep_pause.value()
        sweep_count = self.main_window.sm_sweep_count.value() if not quick_measurement else 1

        return SweepParameters(
            sweep_souce,
            sweep_start,
            sweep_step,
            sweep_end,
            sweep_compliance,
            constant_source,
            constant_output,
            constant_compliance,
            measurement_pause,
            sweep_pause,
            sweep_count,
        )

    def get_datastream_parameters(self):
        smu_1_souce = Source.VOLTAGE if self.main_window.ds_voltage_1 else Source.CURRENT
        smu_1_output = self.main_window.ds_output_1.value()
        smu_1_compliance = self.main_window.ds_compliance_1.value()

        smu_2_source = Source.VOLTAGE if self.main_window.ds_voltage_2 else Source.CURRENT
        smu_2_output = self.main_window.ds_output_2.value()
        smu_2_compliance = self.main_window.ds_compliance_2.value()

        measurement_duration = -1
        measurement_count = -1

        if self.main_window.ds_continuous.isChecked():
            measurement_mode = DatastreamMode.CONTINUOUS
        elif self.main_window.ds_fixed_duration.isChecked():
            measurement_mode = DatastreamMode.FIXED_DURATION
            measurement_duration = self.main_window.ds_duration.value()
        else:
            measurement_mode = DatastreamMode.FIXED_COUNT
            measurement_count = self.main_window.ds_num_measurements.value()

        measurement_pause = self.main_window.ds_measurement_pause.value()

        return DatastreamParameters(
            smu_1_souce,
            smu_1_output,
            smu_1_compliance,
            smu_2_source,
            smu_2_output,
            smu_2_compliance,
            measurement_mode,
            measurement_pause,
            measurement_duration,
            measurement_count,
        )

    def sm_get_smu_1_source(self):
        return Source.VOLTAGE if self.main_window.sm_voltage_1.isChecked() else Source.CURRENT

    def sm_get_smu_2_source(self):
        return Source.VOLTAGE if self.main_window.sm_voltage_2.isChecked() else Source.CURRENT

    def ds_get_smu_1_source(self):
        return Source.VOLTAGE if self.main_window.ds_voltage_1.isChecked() else Source.CURRENT

    def ds_get_smu_2_source(self):
        return Source.VOLTAGE if self.main_window.ds_voltage_2.isChecked() else Source.CURRENT

    def ds_get_smu_1(self):
        return self.get_smu_from_serial(self.main_window.smu_select_ds_1.currentText(), True)

    def ds_get_smu_2(self):
        return self.get_smu_from_serial(self.main_window.smu_select_ds_2.currentText(), True)

    def sm_get_smu_1(self):
        return self.get_smu_from_serial(self.main_window.smu_select_sm_1.currentText(), True)

    def sm_get_smu_2(self):
        return self.get_smu_from_serial(self.main_window.smu_select_sm_2.currentText(), True)

    def sm_get_sweep_smu(self):
        if self.main_window.sm_sweep_1.isChecked():
            return self.sm_get_smu_1()
        else:
            return self.sm_get_smu_2()

    def sm_get_constant_smu(self):
        if not self.main_window.sm_sweep_1.isChecked():
            return self.sm_get_smu_1()
        else:
            return self.sm_get_smu_2()

    @Slot()
    def sm_update_sources(self):

        if self.main_window.sm_sweep_1.isChecked():
            sweep_source = self.sm_get_smu_1_source()
            constant_source = self.sm_get_smu_2_source()
        else:
            sweep_source = self.sm_get_smu_2_source()
            constant_source = self.sm_get_smu_1_source()

        set_input_mode(self.main_window.sm_sweep_start, False, sweep_source)
        set_input_mode(self.main_window.sm_sweep_step, True, sweep_source)
        set_input_mode(self.main_window.sm_sweep_end, False, sweep_source)
        set_input_mode(self.main_window.sm_sweep_compliance, False, Source.CURRENT if sweep_source is Source.VOLTAGE else Source.VOLTAGE)
        set_input_mode(self.main_window.sm_quick_measurement_sweep_step, True, sweep_source)

        set_input_mode(self.main_window.sm_constant_output, False, constant_source)
        set_input_mode(self.main_window.sm_constant_compliance, False, Source.CURRENT if constant_source is Source.VOLTAGE else Source.VOLTAGE)

    @Slot()
    def ds_update_sources(self):
        set_input_mode(self.main_window.ds_output_1, False, self.ds_get_smu_1_source())
        set_input_mode(self.main_window.ds_compliance_1, False, Source.CURRENT if self.ds_get_smu_1_source() is Source.VOLTAGE else Source.VOLTAGE)

        set_input_mode(self.main_window.ds_output_2, False, self.ds_get_smu_2_source())
        set_input_mode(self.main_window.ds_compliance_2, False, Source.CURRENT if self.ds_get_smu_2_source() is Source.VOLTAGE else Source.VOLTAGE)

    @Slot()
    def ds_update_mode(self):
        self.main_window.ds_num_measurements.setDisabled(True)
        self.main_window.ds_num_measurements_label.setDisabled(True)
        self.main_window.ds_duration.setDisabled(True)
        self.main_window.ds_duration_label.setDisabled(True)

        if self.main_window.ds_fixed_duration.isChecked():
            self.main_window.ds_duration.setDisabled(False)
            self.main_window.ds_duration_label.setDisabled(False)
        elif self.main_window.ds_fixed_num.isChecked():
            self.main_window.ds_num_measurements.setDisabled(False)
            self.main_window.ds_num_measurements_label.setDisabled(False)

    @Slot()
    def ds_add_data(self, mp: MeasurementPoint):
        self.ds_last_run = pd.concat(
            (
                self.ds_last_run,
                pd.DataFrame(
                    {
                        "smu_1_voltage": [mp.smu_1_voltage],
                        "smu_1_current": [mp.smu_1_current],
                        "smu_2_voltage": [mp.smu_2_voltage],
                        "smu_2_current": [mp.smu_2_current],
                        "time": [mp.time],
                    }
                ),
            )
        )

        self.main_window.ds_plot.update_data(self.ds_last_run)

    @Slot()
    def sm_add_data(self, mp: MeasurementPoint):
        self.sm_last_run[-1] = pd.concat(
            (
                self.sm_last_run[-1],
                pd.DataFrame(
                    {
                        "smu_1_voltage": [mp.smu_1_voltage],
                        "smu_1_current": [mp.smu_1_current],
                        "smu_2_voltage": [mp.smu_2_voltage],
                        "smu_2_current": [mp.smu_2_current],
                        "time": [mp.time],
                    }
                ),
            )
        )

        self.main_window.sm_plot.update_data(self.sm_last_run[-1])

    def disable_all(self, exclude_widgets: List[QWidget] = None):
        # Note the state of the disabled
        self._widget_enabled_state = {}

        for c in self.main_window.findChildren(QWidget):
            c: QWidget

            self._widget_enabled_state[c] = c.isEnabled()

            c.setEnabled(False)

        # Reenable widgets (and their parents) we want to exclude from being disabled
        # If the parent is disable, the child is also disabled
        for widget in exclude_widgets:
            while widget:
                widget.setEnabled(True)
                widget = widget.parentWidget()

    def restore_all_enable_states(self):
        for widget, enabled in self._widget_enabled_state.items():
            widget.setEnabled(enabled)

    @Slot()
    def sm_run_measurement(self, quick_measurement=False):

        # Show a confirmation if we are about to overwrite data
        if len(self.sm_last_run[0]) > 0:

            confirmation_box = QMessageBox(
                QMessageBox.Icon.Warning,
                "Warning: Data will be discarded",
                "Continuing will discard data from previous run, are you sure you want to continue?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
            )
            confirmation_box.setDefaultButton(QMessageBox.StandardButton.Cancel)

            if confirmation_box.exec() == QMessageBox.StandardButton.Cancel:
                return

        self.sm_last_run = [initialize_data()]

        # self.disable_all([...])

        params = self.get_sweep_parameters(quick_measurement)

        self._thread_sweep = SweepMeasurementWorker(params, self.sm_get_sweep_smu(), self.sm_get_constant_smu())
        self._thread_sweep.finished.connect(self.sm_stop_streaming)
        self._thread_sweep.measurement_made.connect(self.sm_add_data)

        # Make sure that when a sweep finshes, we create a new column
        self._thread_sweep.sweep_complete.connect(lambda: self.sm_last_run.append(initialize_data()))

        # Disable the measurement buttons
        self.main_window.sm_quick_measurement_run.setEnabled(False)
        self.main_window.sm_measurement_run.setEnabled(False)

        self._thread_sweep.start()

        # Enable the stop measurement button and connect it to our thread
        self.main_window.sm_stop.clicked.connect(self._thread_sweep.stop)
        self.main_window.sm_stop.setEnabled(True)

    @Slot()
    def sm_stop_streaming(self):

        self.main_window.sm_quick_measurement_run.setEnabled(True)
        self.main_window.sm_measurement_run.setEnabled(True)

        # Disable the stop measurement button
        self.main_window.sm_stop.setEnabled(False)

        self.main_window.sm_save_data.setEnabled(True)

    @Slot()
    def sm_quick_measurement_run_clicked(self):
        pass

    @Slot()
    def ds_stream_clicked(self):

        # Show a confirmation if we are about to overwrite data
        if len(self.ds_last_run) > 0:

            confirmation_box = QMessageBox(
                QMessageBox.Icon.Warning,
                "Warning: Data will be discarded",
                "Continuing will discard data from previous run, are you sure you want to continue?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
            )
            confirmation_box.setDefaultButton(QMessageBox.StandardButton.Cancel)

            if confirmation_box.exec() == QMessageBox.StandardButton.Cancel:
                return

        self.ds_last_run = initialize_data()

        # self.disable_all([self.main_window.ds_stream, self.main_window.ds_plot_settings])

        params = self.get_datastream_parameters()

        self._thread_datastream = DatastreamMeasurementWorker(params, self.ds_get_smu_1(), self.ds_get_smu_2())
        self._thread_datastream.finished.connect(self.ds_stop_streaming)
        self._thread_datastream.measurement_made.connect(self.ds_add_data)

        # Alter stream button connectins and text
        self.main_window.ds_stream.clicked.disconnect()
        self.main_window.ds_stream.clicked.connect(self._thread_datastream.stop)
        self.main_window.ds_stream.setText("Stop streaming")

        self._thread_datastream.start()

    @Slot()
    def ds_stop_streaming(self):
        self.main_window.ds_stream.clicked.disconnect()
        self.main_window.ds_stream.clicked.connect(self.ds_stream_clicked)
        self.main_window.ds_stream.setText("Start streaming")

        self.main_window.ds_save_data.setEnabled(True)

        # self.restore_all_enable_states()

    @Slot()
    def ds_choose_data_file_path(self):
        path, _ = QFileDialog(self.main_window).getSaveFileName(self.main_window, filter="*.xlsx")

        if path != "":
            self.ds_data_file_path = path
            self.main_window.ds_data_path.setText(self.ds_data_file_path)

    @Slot()
    def sm_choose_data_file_path(self):
        path, _ = QFileDialog(self.main_window).getSaveFileName(self.main_window, filter="*.xlsx")

        if path != "":
            self.sm_data_file_path = path
            self.main_window.sm_data_path.setText(self.sm_data_file_path)

    @Slot()
    def ds_save_last_run(self):

        # If we have not chosen a file path, choose one now
        if self.ds_data_file_path == "":
            self.ds_choose_data_file_path()

            # If we rejected choosing a file path, just return at this point
            if self.ds_data_file_path == "":
                return

        # === Construct the metadata information ===

        # Check to see if we have an excel file already. If so, load it,
        # otherwise create it
        if os.path.exists(self.ds_data_file_path):
            wb = load_workbook(self.ds_data_file_path)
        else:
            wb = Workbook()

        # First sheet should be called metadata
        ws = wb.active
        ws.title = "Metadata"

        light_dark = ""

        if self.main_window.ds_light.isChecked():
            light_dark = "light"
        elif self.main_window.ds_dark.isChecked():
            light_dark = "dark"

        ws["A1"] = f"Wafer #: {self.main_window.ds_wafer_num.text()}"
        ws["A2"] = f"Chip #: {self.main_window.ds_chip_num.text()}"
        ws["A3"] = f"Step of process: {self.main_window.ds_step_of_process.text()}"
        ws["A4"] = f"Light/dark: {light_dark}"
        ws["A5"] = f"Date & time: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}"
        ws["A6"] = f"Chip #: {self.main_window.ds_chip_num.text()}"
        ws["A7"] = f"Comments: {self.main_window.ds_comments.toPlainText()}"

        wb.save(self.ds_data_file_path)
        wb.close()

        # Each measurement has 2 sheets associated with it, data from SMU1 and
        # data from SMU2. Floor divide to find out which measurement we are on.
        # We subtract 1 to account for the metadata sheet
        measurement_number = (len(wb.sheetnames) - 1) // 2

        # We need to reset the indexes
        self.ds_last_run = self.ds_last_run.reset_index(drop=True)

        with pd.ExcelWriter(self.ds_data_file_path, mode="a", if_sheet_exists="replace") as writer:
            self.ds_last_run[["time", "smu_1_voltage", "smu_1_current"]].to_excel(writer, index=False, sheet_name=f"Measurement {measurement_number} - SMU 1")
            self.ds_last_run[["time", "smu_2_voltage", "smu_2_current"]].to_excel(writer, index=False, sheet_name=f"Measurement {measurement_number} - SMU 2")

        # Reset last run data
        self.ds_last_run = initialize_data()
        self.main_window.ds_save_data.setEnabled(False)

    @Slot()
    def sm_save_last_run(self):

        # If we have not chosen a file path, choose one now
        if self.sm_data_file_path == "":
            self.sm_choose_data_file_path()

            # If we rejected choosing a file path, just return at this point
            if self.sm_data_file_path == "":
                return

        # === Construct the metadata information ===

        # Check to see if we have an excel file already. If so, load it,
        # otherwise create it
        if os.path.exists(self.sm_data_file_path):
            wb = load_workbook(self.sm_data_file_path)
        else:
            wb = Workbook()

        # First sheet should be called metadata
        ws = wb.active
        ws.title = "Metadata"

        light_dark = ""

        if self.main_window.sm_light.isChecked():
            light_dark = "light"
        elif self.main_window.sm_dark.isChecked():
            light_dark = "dark"

        ws["A1"] = f"Wafer #: {self.main_window.sm_wafer_num.text()}"
        ws["A2"] = f"Chip #: {self.main_window.sm_chip_num.text()}"
        ws["A3"] = f"Step of process: {self.main_window.sm_step_of_process.text()}"
        ws["A4"] = f"Light/dark: {light_dark}"
        ws["A5"] = f"Date & time: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}"
        ws["A6"] = f"Chip #: {self.main_window.sm_chip_num.text()}"
        ws["A7"] = f"Comments: {self.main_window.sm_comments.toPlainText()}"

        wb.save(self.sm_data_file_path)
        wb.close()

        # Each measurement has 2 sheets associated with it, data from SMU1 and
        # data from SMU2. Floor divide to find out which measurement we are on.
        # We subtract 1 to account for the metadata sheet
        measurement_number = (len(wb.sheetnames) - 1) // 2

        with pd.ExcelWriter(self.sm_data_file_path, mode="a", if_sheet_exists="replace") as writer:
            # The difference with a sweep measurement is that it can have more than
            # one sweep run, so we need to append all of these runs to the data
            for sweep in self.sm_last_run:

                # We need to reset the indexes so we can join later
                sweep = sweep.reset_index(drop=True)

                # Now write the data to the excel sheet
                sweep[["time", "smu_1_voltage", "smu_1_current"]].to_excel(writer, index=False, sheet_name=f"Measurement {measurement_number} - SMU 1")
                sweep[["time", "smu_2_voltage", "smu_2_current"]].to_excel(writer, index=False, sheet_name=f"Measurement {measurement_number} - SMU 2")

                measurement_number += 1

        # Reset last run data
        self.sm_last_run = [initialize_data()]
        self.main_window.sm_save_data.setEnabled(False)


def initialize_data():
    df = pd.DataFrame(columns=["smu_1_voltage", "smu_1_current", "smu_2_voltage", "smu_2_current", "time"])

    for column in df.columns:
        df[column] = df[column].astype(float)

    return df


def set_input_mode(input: QDoubleSpinBox, is_step: bool, source: Source):

    if source is Source.VOLTAGE:
        input.setMaximum(21)
        input.setMinimum(-21)
        input.setSuffix(" V")
    else:
        input.setMaximum(1.05)
        input.setMinimum(-1.05)
        input.setSuffix(" A")

    if is_step:
        input.setMinimum(0.000001)


if __name__ == "__main__":

    import sys

    from PySide6.QtWidgets import QApplication

    # Create the application instance
    app = QApplication(sys.argv)

    controller = Controller()
    controller.smu_search_clicked()

    # Enter the main event loop
    sys.exit(app.exec())
